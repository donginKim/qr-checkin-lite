#!/usr/bin/env python3
"""
엑셀 명단 변환 스크립트

입력 엑셀 형식:
- 시트 이름 = 구역 이름
- C열: 이름
- D열: 세례명
- E열: 전화번호
- 4번째 줄부터 데이터 시작

출력 엑셀 형식 (시스템 업로드용):
- A열: 이름
- B열: 전화번호
- C열: 세례명
- D열: 구역

사용법:
    python convert_excel.py 원본파일.xlsx [출력파일.xlsx]
    
    출력파일명을 지정하지 않으면 '변환_원본파일.xlsx'로 저장됩니다.

필요 패키지:
    pip install openpyxl
"""

import sys
import os
import re
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
except ImportError:
    print("❌ openpyxl 패키지가 필요합니다.")
    print("   설치: pip install openpyxl")
    sys.exit(1)


def clean_district_name(name: str) -> str:
    """구역명에서 앞의 숫자와 기호 제거
    
    예시:
      "1. 본당구역" -> "본당구역"
      "2.청년구역" -> "청년구역"
      "3) 장년구역" -> "장년구역"
      "4 - 교우구역" -> "교우구역"
    """
    # 앞의 숫자, 점, 괄호, 하이픈, 공백 제거
    cleaned = re.sub(r'^[\d\s.\-)\]]+', '', name)
    return cleaned.strip()


def convert_excel(input_file: str, output_file: str = None):
    """엑셀 파일을 시스템 업로드 형식으로 변환"""
    
    if not os.path.exists(input_file):
        print(f"❌ 파일을 찾을 수 없습니다: {input_file}")
        sys.exit(1)
    
    # 출력 파일명 설정
    if output_file is None:
        input_path = Path(input_file)
        output_file = str(input_path.parent / f"변환_{input_path.name}")
    
    print(f"📂 원본 파일: {input_file}")
    print(f"📝 출력 파일: {output_file}")
    print()
    
    # 원본 파일 로드
    wb_input = load_workbook(input_file)
    
    # 새 워크북 생성
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = "회원 명단"
    
    # 헤더 스타일
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 작성
    headers = ["이름", "전화번호", "세례명", "구역"]
    for col, header in enumerate(headers, 1):
        cell = ws_output.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    output_row = 2
    total_count = 0
    stats = {}  # 구역별 통계
    
    # 모든 시트 처리
    for sheet_name in wb_input.sheetnames:
        ws_input = wb_input[sheet_name]
        district = clean_district_name(sheet_name)  # 시트 이름에서 숫자 제거
        
        print(f"📋 시트 처리 중: {sheet_name} → 구역명: {district}")
        
        sheet_count = 0
        
        # 4번째 줄부터 데이터 읽기
        for row in range(4, ws_input.max_row + 1):
            # C열: 이름, D열: 세례명, E열: 전화번호
            name = ws_input.cell(row=row, column=3).value  # C열
            baptismal_name = ws_input.cell(row=row, column=4).value  # D열
            phone = ws_input.cell(row=row, column=5).value  # E열
            
            # 이름이 비어있으면 스킵
            if not name or str(name).strip() == "":
                continue
            
            # 데이터 정리
            name = str(name).strip()
            baptismal_name = str(baptismal_name).strip() if baptismal_name else ""
            phone = str(phone).strip() if phone else ""
            
            # 전화번호 형식 정리 (숫자만 추출 후 하이픈 추가)
            phone_digits = ''.join(filter(str.isdigit, phone))
            if len(phone_digits) == 11:
                phone = f"{phone_digits[:3]}-{phone_digits[3:7]}-{phone_digits[7:]}"
            elif len(phone_digits) == 10:
                phone = f"{phone_digits[:3]}-{phone_digits[3:6]}-{phone_digits[6:]}"
            
            # 출력 파일에 쓰기
            ws_output.cell(row=output_row, column=1, value=name)  # A: 이름
            ws_output.cell(row=output_row, column=2, value=phone)  # B: 전화번호
            ws_output.cell(row=output_row, column=3, value=baptismal_name)  # C: 세례명
            ws_output.cell(row=output_row, column=4, value=district)  # D: 구역
            
            output_row += 1
            sheet_count += 1
            total_count += 1
        
        if sheet_count > 0:
            stats[district] = sheet_count
            print(f"   ✅ {sheet_count}명 변환됨")
        else:
            print(f"   ⚠️ 데이터 없음")
    
    # 열 너비 조정
    ws_output.column_dimensions['A'].width = 15  # 이름
    ws_output.column_dimensions['B'].width = 18  # 전화번호
    ws_output.column_dimensions['C'].width = 15  # 세례명
    ws_output.column_dimensions['D'].width = 12  # 구역
    
    # 저장
    wb_output.save(output_file)
    
    print()
    print("=" * 50)
    print(f"✅ 변환 완료!")
    print(f"📊 총 {total_count}명의 회원 데이터가 변환되었습니다.")
    print()
    print("📈 구역별 현황:")
    for district, count in sorted(stats.items()):
        print(f"   - {district}: {count}명")
    print()
    print(f"💾 출력 파일: {output_file}")
    print()
    print("🚀 이제 관리자 페이지에서 이 파일을 업로드하세요!")


def main():
    if len(sys.argv) < 2:
        print("사용법: python convert_excel.py 원본파일.xlsx [출력파일.xlsx]")
        print()
        print("예시:")
        print("  python convert_excel.py 명단.xlsx")
        print("  python convert_excel.py 명단.xlsx 업로드용_명단.xlsx")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    convert_excel(input_file, output_file)


if __name__ == "__main__":
    main()
